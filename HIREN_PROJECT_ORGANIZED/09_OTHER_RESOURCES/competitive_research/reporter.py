"""
Excel Reporter: builds formatted Excel workbooks with multiple sheets
for each website's competitive research data.
"""

import os
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# Colour palette
COLOR_HEADER = "1F4E79"       # dark blue
COLOR_SUBHEADER = "2E75B6"    # medium blue
COLOR_ALT_ROW = "D6E4F7"     # light blue
COLOR_HIGHLIGHT = "FFF2CC"    # yellow highlight
COLOR_GREEN = "E2EFDA"        # light green
COLOR_ORANGE = "FCE4D6"       # light orange
COLOR_WHITE = "FFFFFF"


def _header_font(white: bool = True) -> Font:
    return Font(name="Calibri", bold=True, color="FFFFFF" if white else "000000", size=11)


def _cell_font() -> Font:
    return Font(name="Calibri", size=10)


def _header_fill(color: str = COLOR_HEADER) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_ALT_ROW)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_column_widths(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def _write_header_row(ws, headers: List[str], row: int = 1, color: str = COLOR_HEADER):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = _header_font()
        cell.fill = _header_fill(color)
        cell.alignment = _center()
        cell.border = _thin_border()


def _write_data_rows(ws, data: List[List[Any]], start_row: int = 2):
    for row_idx, row_data in enumerate(data, start_row):
        fill = _alt_fill() if row_idx % 2 == 0 else PatternFill("solid", fgColor=COLOR_WHITE)
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _cell_font()
            cell.fill = fill
            cell.alignment = _left()
            cell.border = _thin_border()


def write_competitors_sheet(ws, competitors: List[Dict]):
    ws.title = "Competitors"
    headers = [
        "Rank", "Business Name", "Website / Domain", "Snippet / Description",
        "Location Found", "Source Query",
    ]
    _write_header_row(ws, headers)
    data = []
    for c in competitors:
        data.append([
            c.get("rank", ""),
            c.get("name", ""),
            c.get("url", ""),
            c.get("snippet", "")[:200],
            "Long Island, NY area",
            c.get("source_query", ""),
        ])
    _write_data_rows(ws, data)
    ws.freeze_panes = "A2"
    _auto_column_widths(ws)


def write_products_sheet(ws, products: List[Dict]):
    ws.title = "Products & Pricing"
    headers = [
        "Competitor", "Product Name", "Price",
        "Description", "Image File Saved", "Product URL",
    ]
    _write_header_row(ws, headers, color=COLOR_SUBHEADER)
    data = []
    for p in products:
        data.append([
            p.get("competitor", ""),
            p.get("name", ""),
            p.get("price", ""),
            p.get("description", "")[:150],
            p.get("saved_image", ""),
            p.get("product_url", ""),
        ])
    _write_data_rows(ws, data)
    ws.freeze_panes = "A2"
    _auto_column_widths(ws)


def write_contacts_sheet(ws, contacts: List[Dict]):
    ws.title = "Business Contacts"
    headers = [
        "Competitor", "Domain", "Email (Public)",
        "Phone (Public)", "Address", "Contact Page URL",
    ]
    _write_header_row(ws, headers, color="145A32")
    data = []
    for c in contacts:
        data.append([
            c.get("competitor", ""),
            c.get("domain", ""),
            c.get("email", ""),
            c.get("phone", ""),
            c.get("address", ""),
            c.get("contact_page", ""),
        ])
    _write_data_rows(ws, data)
    ws.freeze_panes = "A2"
    _auto_column_widths(ws)


def write_images_sheet(ws, image_records: List[Dict]):
    ws.title = "Images Downloaded"
    headers = [
        "Competitor", "Image URL", "Saved Filename", "Status",
    ]
    _write_header_row(ws, headers, color="7B241C")
    data = []
    for r in image_records:
        data.append([
            r.get("competitor", ""),
            r.get("image_url", ""),
            r.get("saved_filename", ""),
            r.get("status", ""),
        ])
    _write_data_rows(ws, data)
    ws.freeze_panes = "A2"
    _auto_column_widths(ws)


def write_price_analysis_sheet(ws, products: List[Dict], site_config: Dict):
    """Generate a pricing analysis summary with our suggested prices."""
    ws.title = "Price Analysis"

    # Title
    ws.cell(row=1, column=1).value = f"Price Analysis — {site_config['business_name']}"
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14, color=COLOR_HEADER)
    ws.merge_cells("A1:G1")

    headers = [
        "Category", "Competitor Prices Found",
        "Min Price", "Max Price", "Avg Price",
        "Suggested Our Price", "Pricing Strategy",
    ]
    _write_header_row(ws, headers, row=2, color=COLOR_HEADER)

    # Group products by category (approximate by keyword matching)
    categories = site_config.get("product_categories", [])
    price_range = site_config.get("price_range", {"min": 5, "max": 500})

    rows = []
    for cat in categories:
        cat_prices = []
        for p in products:
            name = (p.get("name", "") + " " + p.get("description", "")).lower()
            cat_key = cat.lower().replace("&", "").replace(" ", "")
            name_key = name.replace("&", "").replace(" ", "")
            price_str = p.get("price", "")
            if price_str:
                try:
                    val = float(price_str.replace("$", "").replace(",", ""))
                    cat_prices.append(val)
                except Exception:
                    pass

        if cat_prices:
            min_p = min(cat_prices)
            max_p = max(cat_prices)
            avg_p = sum(cat_prices) / len(cat_prices)
            our_price = round(avg_p * 0.95, 2)
            strategy = "Match avg, undercut by 5%"
            prices_found = ", ".join(f"${v:.2f}" for v in sorted(set(cat_prices))[:5])
        else:
            min_p = price_range["min"]
            max_p = price_range["max"]
            avg_p = (min_p + max_p) / 2
            our_price = round(avg_p * 0.90, 2)
            strategy = "No comp data — use range midpoint"
            prices_found = "Not scraped"

        rows.append([
            cat,
            prices_found,
            f"${min_p:.2f}",
            f"${max_p:.2f}",
            f"${avg_p:.2f}",
            f"${our_price:.2f}",
            strategy,
        ])

    _write_data_rows(ws, rows, start_row=3)
    ws.freeze_panes = "A3"
    _auto_column_widths(ws)


def write_market_summary_sheet(ws, site_config: Dict, competitors: List[Dict],
                                products: List[Dict], contacts: List[Dict]):
    ws.title = "Market Summary"

    def add_section_title(row, title, color=COLOR_HEADER):
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        ws.merge_cells(f"A{row}:C{row}")
        return row + 1

    def add_kv(row, key, val):
        ws.cell(row=row, column=1, value=key).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=row, column=2, value=val).font = Font(name="Calibri", size=10)
        return row + 1

    row = 1
    ws.cell(row=row, column=1).value = f"Market Intelligence Summary — {site_config['business_name']}"
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=16, color=COLOR_HEADER)
    ws.merge_cells(f"A{row}:D{row}")
    row += 2

    row = add_section_title(row, "Business Profile")
    row = add_kv(row, "Our Website", site_config["domain"])
    row = add_kv(row, "Niche", site_config["niche"])
    row = add_kv(row, "Location", site_config["location"])
    row = add_kv(row, "Categories", len(site_config.get("product_categories", [])))
    row += 1

    row = add_section_title(row, "Competitor Intelligence", COLOR_SUBHEADER)
    row = add_kv(row, "Total Competitors Found", len(competitors))
    if competitors:
        row = add_kv(row, "Top Competitor", competitors[0].get("name", ""))
    row = add_kv(row, "Products Scraped", len(products))
    row = add_kv(row, "Contacts Extracted", sum(1 for c in contacts if c.get("email") or c.get("phone")))
    row += 1

    priced = [p for p in products if p.get("price")]
    if priced:
        row = add_section_title(row, "Pricing Overview", "145A32")
        prices = []
        for p in priced:
            try:
                val = float(p["price"].replace("$", "").replace(",", ""))
                prices.append(val)
            except Exception:
                pass
        if prices:
            row = add_kv(row, "Lowest Price Seen", f"${min(prices):.2f}")
            row = add_kv(row, "Highest Price Seen", f"${max(prices):.2f}")
            row = add_kv(row, "Average Price", f"${sum(prices)/len(prices):.2f}")
            row = add_kv(row, "Price Range", f"${min(prices):.2f} – ${max(prices):.2f}")
        row += 1

    row = add_section_title(row, "Action Items", "7E5109")
    action_items = [
        "Review competitors_analysis.xlsx → Competitors tab for top 10 list",
        "Set product prices based on Price Analysis tab",
        "Download and review images from images/ folder for design inspiration",
        "Use contacts from Business Contacts tab for outreach/partnerships",
        "Update Odoo product catalog with recommended prices",
    ]
    for item in action_items:
        ws.cell(row=row, column=1, value=f"→ {item}").font = Font(name="Calibri", size=10)
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def build_excel_report(
    output_path: Path,
    site_config: Dict,
    competitors: List[Dict],
    products: List[Dict],
    contacts: List[Dict],
    image_records: List[Dict],
):
    """Create the full Excel workbook for one website."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Sheet 1: Market Summary
    ws_summary = wb.create_sheet("Market Summary")
    write_market_summary_sheet(ws_summary, site_config, competitors, products, contacts)

    # Sheet 2: Competitors
    ws_comp = wb.create_sheet("Competitors")
    write_competitors_sheet(ws_comp, competitors)

    # Sheet 3: Products & Pricing
    ws_prod = wb.create_sheet("Products & Pricing")
    write_products_sheet(ws_prod, products)

    # Sheet 4: Price Analysis
    ws_price = wb.create_sheet("Price Analysis")
    write_price_analysis_sheet(ws_price, products, site_config)

    # Sheet 5: Business Contacts
    ws_contacts = wb.create_sheet("Business Contacts")
    write_contacts_sheet(ws_contacts, contacts)

    # Sheet 6: Images
    ws_images = wb.create_sheet("Images Downloaded")
    write_images_sheet(ws_images, image_records)

    wb.save(output_path)
    print(f"  Excel saved: {output_path.name}")
