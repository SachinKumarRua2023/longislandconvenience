"""
Master Celebration & Special Days Excel Builder
LI Gift Basket — Plainview, Long Island, NY

Generates a fully formatted Excel workbook with:
  Sheet 1:  MASTER LIST — all 200+ days with auto-greeting content
  Sheet 2:  GRADUATION & CEREMONIES
  Sheet 3:  MAJOR US HOLIDAYS
  Sheet 4:  PERSONAL MILESTONES (birthday, anniversary, baby, wedding)
  Sheet 5:  RELIGIOUS & CULTURAL
  Sheet 6:  CORPORATE & PROFESSIONAL
  Sheet 7:  SEASONAL & FUN DAYS
  Sheet 8:  LONG ISLAND SPECIAL
  Sheet 9:  7-DAY ALERT (due this week)
  Sheet 10: AUTO-GREETING SCRIPTS (email + WhatsApp ready to copy)
"""

import sys
from pathlib import Path
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBar, FormatObject
from openpyxl.styles.numbers import FORMAT_TEXT

OUTPUT_DIR = Path(r"c:\Users\Sachin Kumar\OneDrive\Desktop\HirenTask\ScrappedMaterial\03_LIGiftBasket")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY = date.today()
YEAR = 2026

# ── Colour palette ────────────────────────────────────────────────────────────
C = {
    "dark_purple":  "4A235A",
    "purple":       "7D3C98",
    "light_purple": "D2B4DE",
    "pink":         "F1948A",
    "dark_blue":    "1B4F72",
    "blue":         "2980B9",
    "light_blue":   "AED6F1",
    "dark_green":   "1E8449",
    "green":        "27AE60",
    "light_green":  "A9DFBF",
    "gold":         "B7950B",
    "yellow":       "F9E79F",
    "orange":       "E67E22",
    "light_orange": "FAD7A0",
    "red":          "C0392B",
    "light_red":    "F1948A",
    "dark_gray":    "2C3E50",
    "gray":         "85929E",
    "light_gray":   "EBF5FB",
    "white":        "FFFFFF",
    "teal":         "0E6655",
    "maroon":       "78281F",
}

CATEGORY_COLORS = {
    "Graduation & Ceremony":    (C["dark_purple"], C["light_purple"]),
    "Major US Holiday":         (C["dark_blue"],   C["light_blue"]),
    "Personal Milestone":       (C["red"],          C["light_red"]),
    "Religious":                (C["gold"],         C["yellow"]),
    "Cultural":                 (C["teal"],         C["light_green"]),
    "Corporate & Professional": (C["dark_gray"],    C["light_gray"]),
    "Seasonal":                 (C["dark_green"],   C["light_green"]),
    "Long Island Special":      (C["orange"],       C["light_orange"]),
    "Food & Fun":               (C["purple"],       C["light_purple"]),
    "Health & Awareness":       (C["blue"],         C["light_blue"]),
    "Love & Romance":           (C["red"],          C["light_red"]),
    "Shopping & Sale":          (C["maroon"],       C["yellow"]),
}


def _font(bold=False, size=10, color="000000", name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


# ── ALL SPECIAL DAYS DATA ─────────────────────────────────────────────────────
def build_all_days():
    today = TODAY

    def d(m, day, name, cat, emoji, disc, basket_idea, email_subject, email_body, whatsapp):
        try:
            dt = date(YEAR, m, day)
        except ValueError:
            return None
        days_left = (dt - today).days
        status = "TODAY" if days_left == 0 else "UPCOMING" if days_left > 0 else "PAST"
        urgency = "URGENT (≤7d)" if 0 < days_left <= 7 else "SOON (≤30d)" if 0 < days_left <= 30 else status
        promo = name.upper().replace(" ", "").replace("'", "")[:8] + str(YEAR)[-2:]
        return {
            "Date": dt.strftime("%B %d, %Y"),
            "Day": dt.strftime("%A"),
            "Month": dt.strftime("%B"),
            "Event": name,
            "Category": cat,
            "Emoji": emoji,
            "Discount %": disc,
            "Days Until": days_left if days_left >= 0 else "",
            "Status": status,
            "Alert": urgency if status == "UPCOMING" else status,
            "Promo Code": promo,
            "Best Basket Idea": basket_idea,
            "Email Subject": email_subject,
            "Email Body": email_body,
            "WhatsApp Message": whatsapp,
            "_date_obj": dt,
            "_cat": cat,
        }

    def make_email(name, emoji, days_text, disc, basket, promo):
        return (
            f"Dear [Customer Name],\n\n"
            f"{emoji} {name} is {days_text}! Celebrate with a beautiful gift basket from "
            f"LI Gift Basket — Plainview, Long Island's favorite gift shop.\n\n"
            f"🎁 Perfect for this occasion: {basket}\n"
            f"💰 Save {disc}% with code: {promo}\n"
            f"🚚 Free delivery across Nassau & Suffolk County (orders $60+)\n\n"
            f"Shop now: https://ligiftbasket.com\n\n"
            f"Warm regards,\nLI Gift Basket | Plainview, Long Island, NY 11803\n📞 (516) 555-0100"
        )

    def make_wa(name, emoji, days_text, disc, promo):
        return (
            f"{emoji} *{name}* is {days_text}!\n"
            f"Celebrate with a beautiful gift basket 🎁\n"
            f"💰 {disc}% OFF with code *{promo}*\n"
            f"🚚 Free delivery on Long Island\n"
            f"👉 https://ligiftbasket.com"
        )

    rows = []
    def add(m, day, name, cat, emoji, disc, basket):
        try:
            dt = date(YEAR, m, day)
        except ValueError:
            return
        days_left = (dt - TODAY).days
        days_text = "TODAY" if days_left == 0 else f"in {days_left} days" if days_left > 0 else f"{abs(days_left)} days ago"
        promo = name.upper().replace(" ","").replace("'","").replace("&","")[:8] + str(YEAR)[-2:]
        status = "TODAY" if days_left == 0 else "UPCOMING" if days_left > 0 else "PAST"
        urgency = "URGENT (≤7d)" if 0 < days_left <= 7 else "SOON (≤30d)" if 0 < days_left <= 30 else status
        rows.append({
            "Date": dt.strftime("%B %d, %Y"),
            "Day": dt.strftime("%A"),
            "Month": dt.strftime("%B"),
            "Event": name,
            "Category": cat,
            "Emoji": emoji,
            "Discount %": disc,
            "Days Until": days_left if days_left >= 0 else "PAST",
            "Status": status,
            "Alert": urgency,
            "Promo Code": promo,
            "Best Basket Idea": basket,
            "Email Subject": f"{emoji} {name} is {days_text} — {disc}% OFF Gift Baskets!",
            "Email Body": make_email(name, emoji, days_text, disc, basket, promo),
            "WhatsApp Message": make_wa(name, emoji, days_text, disc, promo),
            "_date_obj": dt,
            "_cat": cat,
        })

    # ── GRADUATION & CEREMONY ─────────────────────────────────────────────────
    G = "Graduation & Ceremony"
    add(5, 15, "High School Graduation Season (May)",        G, "🎓", 25, "Graduation Celebration Basket")
    add(5, 22, "College Graduation Season (May)",            G, "🎓", 25, "College Grad Gourmet Basket")
    add(5, 29, "Spring Graduation Ceremonies (LI Schools)",  G, "🎓", 25, "Class of 2026 Gift Basket")
    add(6, 5,  "June Graduation Ceremonies",                 G, "🎓", 25, "Congrats Grad Luxury Basket")
    add(6, 12, "Summer Graduation Parties",                  G, "🎓", 20, "Party Snack & Treat Basket")
    add(8, 14, "Summer School Graduation",                   G, "🎓", 15, "Achievement Celebration Basket")
    add(12, 18, "Winter/December Graduation Ceremonies",     G, "🎓", 25, "Winter Grad Holiday Basket")
    add(12, 21, "Winter Commencement (Universities)",        G, "🎓", 20, "Congrats Class Basket")
    add(5, 16, "Kindergarten Graduation",                    G, "🌟", 15, "Little Graduate Sweet Basket")
    add(6, 13, "Middle School Promotion Ceremony",           G, "🏅", 15, "Moving Up Celebration Basket")
    add(6, 10, "Pre-K & Nursery Graduation",                 G, "👶", 10, "Little Achiever Snack Basket")
    add(5, 23, "Nursing School Pinning Ceremony",            G, "👩‍⚕️",20, "Future Nurse Wellness Basket")
    add(5, 30, "Law School Graduation (JD Ceremony)",        G, "⚖️", 25, "Esquire Gourmet Gift Basket")
    add(5, 17, "Medical School Graduation (MD Ceremony)",    G, "🩺", 25, "Doctor's Celebration Basket")
    add(6, 6,  "MBA Graduation & Business School",           G, "💼", 25, "Executive Success Basket")
    add(6, 20, "Trade & Vocational School Graduation",       G, "🔧", 20, "Skilled Trades Celebration Basket")
    add(5, 14, "Nassau County School Graduation (NCSD)",     G, "📚", 20, "Long Island Graduate Basket")
    add(5, 21, "Suffolk County School Graduation (SCSD)",    G, "📚", 20, "Long Island Scholar Basket")
    add(6, 7,  "Military Graduation / Boot Camp Completion", G, "🎖️",25, "American Hero Gift Basket")

    # ── MAJOR US HOLIDAYS ─────────────────────────────────────────────────────
    H = "Major US Holiday"
    add(1, 1,  "New Year's Day",                 H, "🎆", 15, "New Year Celebration Basket")
    add(2, 14, "Valentine's Day",                H, "❤️", 20, "Chocolate & Roses Love Basket")
    add(3, 17, "St. Patrick's Day",              H, "☘️", 17, "Irish Luck Treat Basket")
    add(4, 5,  "Easter Sunday",                  H, "🐣", 20, "Easter Candy & Treat Basket")
    add(5, 10, "Mother's Day",                   H, "🌹", 25, "Mom's Spa & Pampering Basket")
    add(5, 25, "Memorial Day",                   H, "🎖️",15, "American Hero Gift Basket")
    add(6, 21, "Father's Day",                   H, "👔", 25, "Dad's Gourmet Grilling Basket")
    add(7, 4,  "Independence Day",               H, "🎆", 20, "Red White & Blue Celebration Basket")
    add(9, 7,  "Labor Day",                      H, "👷", 20, "End-of-Summer BBQ Basket")
    add(10, 31,"Halloween",                      H, "🎃", 20, "Spooky Treat & Candy Basket")
    add(11, 26,"Thanksgiving",                   H, "🦃", 20, "Harvest & Gratitude Gift Basket")
    add(11, 27,"Black Friday",                   H, "🛍️",30, "Holiday Shopping Snack Basket")
    add(12, 2, "Cyber Monday",                   H, "💻", 25, "Online Sale Gift Bundle")
    add(12, 25,"Christmas Day",                  H, "🎄", 25, "Christmas Holiday Gift Basket")
    add(12, 31,"New Year's Eve",                 H, "🥂", 20, "Champagne & Celebration Basket")
    add(6, 19, "Juneteenth",                     H, "✊", 15, "Unity & Heritage Gift Basket")
    add(1, 19, "Martin Luther King Jr. Day",     H, "✊", 15, "Community & Hope Basket")
    add(2, 16, "Presidents' Day",                H, "🇺🇸",10, "All-American Snack Basket")

    # ── PERSONAL MILESTONES ───────────────────────────────────────────────────
    P = "Personal Milestone"
    add(1, 10, "New Baby Arrival Gift",          P, "👶", 20, "Welcome Baby Essentials Basket")
    add(2, 10, "Baby Shower",                    P, "🍼", 20, "Baby Shower Pamper Basket")
    add(3, 15, "Wedding Anniversary",            P, "💍", 25, "Anniversary Luxury Wine Basket")
    add(4, 18, "Wedding Day Gift",               P, "👰", 25, "Bridal Bliss Celebration Basket")
    add(5, 3,  "Bridal Shower",                  P, "💐", 20, "Bride-to-Be Spa Basket")
    add(5, 9,  "Bachelorette Party",             P, "🥂", 15, "Girls Night Out Treat Basket")
    add(6, 14, "Retirement Party",               P, "🏖️",25, "Bon Voyage Retirement Basket")
    add(7, 12, "New Home / Housewarming",        P, "🏠", 20, "Housewarming Gourmet Basket")
    add(8, 20, "First Day of School Gift",       P, "🎒", 15, "Back to School Snack Basket")
    add(9, 5,  "Job Promotion Celebration",      P, "💼", 20, "Career Success Gift Basket")
    add(10, 8, "Birthday (General)",             P, "🎂", 15, "Happy Birthday Treat Basket")
    add(11, 3, "Sweet 16 / Quinceañera",         P, "🎀", 20, "Sweet Sixteen Celebration Basket")
    add(12, 5, "Bar/Bat Mitzvah Gift",           P, "✡️", 20, "Mazel Tov Celebration Basket")
    add(4, 12, "Baptism / Christening",          P, "✝️", 15, "Blessing & Joy Gift Basket")
    add(7, 20, "Engagement Announcement",        P, "💍", 20, "She Said Yes! Celebration Basket")
    add(9, 25, "Get Well / Recovery Gift",       P, "🌻", 15, "Feel Better Soon Comfort Basket")
    add(10, 15,"Sympathy / Condolence Gift",     P, "🕊️", 10, "Comfort & Sympathy Basket")
    add(3, 20, "First Day of Spring Birthday",   P, "🌸", 15, "Spring Birthday Flower Basket")
    add(11, 15,"Military Homecoming",            P, "🎖️",20, "Welcome Home Hero Basket")
    add(6, 28, "Quinceanera Season (June)",      P, "🌺", 20, "Quinceañera Fiesta Basket")

    # ── RELIGIOUS & CULTURAL ─────────────────────────────────────────────────
    R = "Religious"
    add(1, 6,  "Three Kings Day / Epiphany",      R, "👑", 15, "Three Kings Holiday Basket")
    add(2, 5,  "Chinese New Year",                R, "🐉", 15, "Lucky Red Envelope Gift Basket")
    add(3, 8,  "Holi Festival of Colors",         R, "🎨", 15, "Colorful Celebration Sweet Basket")
    add(4, 4,  "Good Friday",                     R, "✝️", 10, "Easter Weekend Treat Basket")
    add(4, 22, "Passover (Pesach)",               R, "🕍", 15, "Kosher Passover Gift Basket")
    add(11, 1, "Diwali (Festival of Lights)",     R, "🪔", 20, "Diwali Sweets & Lights Basket")
    add(11, 1, "All Saints Day",                  R, "✝️", 5,  "Blessed Day Gift Basket")
    add(12, 19,"Hanukkah Begins",                 R, "🕎", 20, "Hanukkah Gelt & Treat Basket")
    add(12, 26,"Kwanzaa Begins",                  R, "🕯️",15, "Kwanzaa Heritage Gift Basket")
    add(5, 4,  "Eid al-Adha / Eid Celebration",  R, "🌙", 20, "Eid Mubarak Celebration Basket")
    add(3, 26, "Nowruz (Persian New Year)",       R, "🌷", 15, "New Year Springtime Basket")
    add(10, 2, "Rosh Hashanah (Jewish New Year)", R, "🍎", 20, "Sweet New Year Honey Basket")
    add(10, 11,"Yom Kippur (Day of Atonement)",   R, "✡️", 10, "Break-the-Fast Gift Basket")
    add(11, 12,"Diwali Gift Season",              R, "🪔", 20, "Diwali Luxury Sweet Box")
    add(12, 6, "St. Nicholas Day",                R, "🎅", 15, "St. Nick Holiday Gift Basket")

    # ── CORPORATE & PROFESSIONAL ─────────────────────────────────────────────
    B = "Corporate & Professional"
    add(1, 16, "National Get to Know Your Customers Day",  B, "🤝", 15, "Client Appreciation Basket")
    add(3, 1,  "Employee Appreciation Day",                B, "👏", 20, "Team Champion Gift Basket")
    add(4, 25, "Administrative Professionals Day",         B, "📋", 20, "Office Hero Gift Basket")
    add(5, 6,  "National Nurses Day",                      B, "💊", 20, "Healthcare Hero Basket")
    add(5, 15, "National Teacher Appreciation Day",        B, "🍎", 20, "Teacher Thank You Basket")
    add(6, 14, "National Flag Day (Corporate Gifting)",    B, "🇺🇸",10, "Patriotic Corporate Basket")
    add(7, 1,  "National Postal Worker Day",               B, "📬", 10, "Essential Worker Thank You Basket")
    add(10, 16,"National Boss Day",                        B, "👔", 20, "Boss's Day Gourmet Basket")
    add(11, 19,"International Men's Day (Client Gift)",    B, "👨", 15, "Gentleman's Luxury Basket")
    add(9, 5,  "National Labor Day (Corporate)",           B, "👷", 15, "Hardworking Team Basket")
    add(12, 1, "Corporate Holiday Gifting Season Begins",  B, "🎁", 25, "Executive Holiday Gift Basket")
    add(12, 15,"Year-End Client Appreciation",             B, "🥂", 25, "Year-End Thank You Basket")
    add(5, 12, "National Receptionists Day",               B, "📞", 15, "Front Desk Hero Basket")
    add(8, 5,  "National Wellness Day (Corporate)",        B, "🧘", 15, "Workplace Wellness Basket")
    add(2, 1,  "Financial Advisor Appreciation Day",       B, "💰", 15, "Financial Advisor Gift Basket")
    add(10, 5, "World Teachers Day",                       B, "📚", 20, "Educator Excellence Basket")

    # ── SEASONAL ─────────────────────────────────────────────────────────────
    S = "Seasonal"
    add(3, 20, "First Day of Spring",     S, "🌸", 10, "Spring Blooms Gift Basket")
    add(6, 21, "First Day of Summer",     S, "☀️", 15, "Summer Fun Beach Basket")
    add(9, 22, "First Day of Fall",       S, "🍂", 15, "Autumn Harvest Basket")
    add(12, 21,"First Day of Winter",     S, "❄️", 10, "Winter Wonderland Basket")
    add(5, 1,  "May Day / Spring Fling",  S, "🌷", 10, "Spring Floral Gift Basket")
    add(10, 1, "Fall/Autumn Gift Season", S, "🎃", 15, "Pumpkin Spice & Everything Nice Basket")
    add(8, 31, "End of Summer Bash",      S, "🌊", 15, "Summer Farewell Beach Basket")
    add(12, 10,"Holiday Gift Season Peak",S, "🎄", 25, "Holiday Festive Gift Box")
    add(1, 15, "Mid-Winter Treat Day",    S, "⛄", 10, "Winter Cozy Night In Basket")
    add(4, 15, "Tax Day Stress Relief",   S, "📊", 15, "Tax Day Survival Basket")
    add(8, 7,  "Long Island Summer Peak", S, "🏖️",15, "Long Island Summer Basket")
    add(11, 20,"Pre-Thanksgiving Season", S, "🦃", 20, "Thanksgiving Pre-Game Basket")

    # ── LONG ISLAND SPECIAL ───────────────────────────────────────────────────
    LI = "Long Island Special"
    add(5, 23, "Long Island Ducks Season Opener",         LI, "⚾", 15, "Game Day Snack Basket")
    add(7, 4,  "Hamptons Fourth of July",                 LI, "🎆", 20, "Hamptons Holiday Gift Basket")
    add(8, 7,  "National Lighthouse Day (LI has 21!)",    LI, "🏮", 10, "Long Island Heritage Basket")
    add(9, 1,  "Back to School Long Island (LISD)",       LI, "🎒", 15, "LI Back-to-School Basket")
    add(10, 11,"Columbus Day Parade (NYC/LI)",            LI, "🎊", 10, "Italian Heritage Feast Basket")
    add(11, 7, "NYC Marathon (LI runners)",               LI, "🏃", 15, "Marathon Recovery Basket")
    add(12, 5, "Long Island Winterfest",                  LI, "❄️", 15, "Winterfest Cozy Gift Basket")
    add(6, 14, "Plainview-Old Bethpage School Events",    LI, "📚", 15, "School Community Basket")
    add(5, 30, "Nassau County Memorial Day Parade",       LI, "🇺🇸",15, "Nassau County Hero Basket")
    add(3, 15, "St. Patrick's Day Parade (LI)",           LI, "☘️", 15, "Long Island Irish Basket")
    add(9, 13, "Long Island Wine Country Harvest",        LI, "🍷", 20, "LI Wine Country Harvest Basket")
    add(10, 5, "Long Island Fall Harvest Festival",       LI, "🎃", 15, "LI Fall Harvest Basket")

    # ── HEALTH & AWARENESS ────────────────────────────────────────────────────
    A = "Health & Awareness"
    add(2, 4,  "World Cancer Day",              A, "🎗️",10, "Strength & Hope Wellness Basket")
    add(4, 7,  "World Health Day",              A, "🏥", 10, "Health & Wellness Gift Basket")
    add(4, 22, "Earth Day",                     A, "🌍", 10, "Eco-Friendly Green Basket")
    add(5, 12, "International Nurses Day",      A, "💊", 20, "Nurse Hero Appreciation Basket")
    add(9, 10, "World Suicide Prevention Day",  A, "💙", 5,  "Mental Health Comfort Basket")
    add(10, 10,"World Mental Health Day",       A, "🧠", 15, "Mindfulness & Relaxation Basket")
    add(11, 14,"World Diabetes Day",            A, "💙", 10, "Sugar-Free Snack Basket")
    add(10, 1, "Breast Cancer Awareness Month", A, "🎀", 15, "Pink Ribbon Care Basket")
    add(11, 13,"World Kindness Day",            A, "💛", 15, "Random Acts of Kindness Basket")
    add(9, 21, "World Alzheimer's Day",         A, "💜", 10, "Memory Care Comfort Basket")
    add(4, 2,  "World Autism Awareness Day",    A, "💙", 10, "Sensory-Friendly Gift Basket")
    add(3, 22, "World Water Day (Eco Gift)",    A, "💧", 5,  "Clean & Green Eco Basket")

    # ── LOVE & ROMANCE ────────────────────────────────────────────────────────
    LR = "Love & Romance"
    add(1, 26, "National Spouses Day",          LR, "💑", 15, "Couples Indulgence Basket")
    add(2, 13, "Galentine's Day",               LR, "💕", 15, "Girlfriends Celebration Basket")
    add(2, 14, "Valentine's Day",               LR, "❤️", 20, "Love & Chocolate Romance Basket")
    add(6, 23, "National Pink Day",             LR, "🌸", 10, "Pretty in Pink Gift Basket")
    add(8, 1,  "National Girlfriends Day",      LR, "👭", 15, "Best Friends Forever Basket")
    add(10, 3, "National Boyfriend Day",        LR, "💑", 15, "Him & Hers Couple Gift Basket")
    add(10, 19,"Sweetest Day",                  LR, "🍬", 15, "Sweetest Chocolate Candy Basket")
    add(7, 30, "International Friendship Day",  LR, "👫", 15, "Best Friend Gift Basket")
    add(11, 11,"Singles Day (11/11)",           LR, "🌟", 15, "Treat Yourself Luxury Basket")

    # ── FOOD & FUN ────────────────────────────────────────────────────────────
    F = "Food & Fun"
    add(1, 15, "National Bagel Day",            F, "🥯", 8,  "NY Bagel & Schmear Basket")
    add(2, 9,  "National Pizza Day",            F, "🍕", 8,  "NY Pizza Night Gift Basket")
    add(3, 14, "National Pi Day (Pie Day)",     F, "🥧", 8,  "Sweet Pie Treat Basket")
    add(4, 14, "National Pecan Day",            F, "🥜", 8,  "Nut Lover's Gift Basket")
    add(5, 15, "National Chocolate Chip Day",   F, "🍪", 10, "Cookie & Bake Basket")
    add(5, 25, "National Wine Day",             F, "🍷", 15, "Wine Lover's Gift Basket")
    add(6, 1,  "National Donut Day",            F, "🍩", 8,  "Donut & Coffee Morning Basket")
    add(7, 7,  "National Chocolate Day",        F, "🍫", 12, "Chocolate Lover's Dream Basket")
    add(7, 18, "National Ice Cream Day",        F, "🍦", 10, "Ice Cream Party Kit Basket")
    add(9, 17, "National Apple Day",            F, "🍎", 8,  "Apple Orchard Harvest Basket")
    add(10, 1, "International Coffee Day",      F, "☕", 10, "Coffee Lover's Morning Basket")
    add(10, 28,"National Chocolate Day #2",     F, "🍫", 12, "Double Chocolate Luxury Basket")
    add(12, 4, "National Cookie Day",           F, "🍪", 10, "Holiday Cookie Gift Tin")
    add(11, 25,"National Parfait Day",          F, "🍨", 8,  "Sweet Parfait Treat Basket")
    add(4, 1,  "April Fools Fun Gift",          F, "🤡", 10, "Silly & Sweet Prank Basket")

    # ── SHOPPING & SALE ───────────────────────────────────────────────────────
    SL = "Shopping & Sale"
    add(1, 1,  "New Year Sale",                 SL, "🎆", 20, "New Year New Gifts Basket")
    add(2, 1,  "February Flash Sale",           SL, "❤️", 15, "February Love Gift Basket")
    add(7, 1,  "Mid-Year Sale",                 SL, "☀️", 20, "Summer Savings Gift Basket")
    add(10, 1, "Fall Sale Kickoff",             SL, "🍂", 20, "Fall Favorites Gift Basket")
    add(11, 27,"Black Friday MEGA SALE",        SL, "🛍️",30, "Black Friday Luxury Gift Set")
    add(11, 30,"Cyber Monday Eve Sale",         SL, "💻", 25, "Online Deal Gift Bundle")
    add(12, 12,"12/12 Annual Sale",             SL, "🎁", 20, "12.12 Holiday Gift Box")
    add(12, 26,"Boxing Day Sale",               SL, "📦", 25, "Boxing Day Bargain Basket")

    # Sort by date
    valid = [r for r in rows if r is not None]
    valid.sort(key=lambda x: x["_date_obj"])
    return valid


# ── EXCEL STYLING HELPERS ────────────────────────────────────────────────────
def write_title_row(ws, title, row=1, col_span=15, bg=C["dark_purple"], fg=C["white"], size=14):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _font(bold=True, size=size, color=fg)
    cell.fill = _fill(bg)
    cell.alignment = _align(h="center")
    cell.border = _border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    _set_row_height(ws, row, 30)


def write_header_row(ws, headers, row=2, bg=C["purple"], fg=C["white"]):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _font(bold=True, size=10, color=fg)
        cell.fill = _fill(bg)
        cell.alignment = _align(h="center")
        cell.border = _border()
    _set_row_height(ws, row, 22)


def write_data_row(ws, row_num, data, category=None):
    cat_colors = CATEGORY_COLORS.get(category, (C["dark_gray"], C["light_gray"]))
    alt_fill = _fill(cat_colors[1]) if row_num % 2 == 0 else _fill(C["white"])

    for c, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.font = _font(size=9)
        cell.fill = alt_fill
        cell.alignment = _align(wrap=True)
        cell.border = _border(color="DDDDDD")

    # Color the Status column (col 9 in master)
    _set_row_height(ws, row_num, 60 if any(isinstance(v, str) and len(str(v)) > 100 for v in data) else 18)


def auto_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ── SHEET BUILDERS ─────────────────────────────────────────────────────────
HEADERS_MASTER = [
    "Date", "Day", "Month", "Event / Occasion", "Category",
    "Emoji", "Discount %", "Days Until", "Status", "Alert Level",
    "Promo Code", "Best Basket Idea",
    "Email Subject", "Email Body (Auto-Greeting)", "WhatsApp Message (Auto-Greeting)"
]

HEADERS_COMPACT = [
    "Date", "Day", "Event / Occasion", "Category", "Emoji",
    "Discount %", "Days Until", "Status", "Promo Code",
    "Best Basket Idea", "Auto Email Subject", "WhatsApp Quick Message"
]

MASTER_WIDTHS = {
    1: 15, 2: 12, 3: 12, 4: 32, 5: 22, 6: 6, 7: 10, 8: 10,
    9: 12, 10: 16, 11: 16, 12: 30, 13: 40, 14: 60, 15: 60
}


def build_master_sheet(ws, rows):
    write_title_row(ws, "LI Gift Basket — Complete 2026 Celebration & Special Days Calendar (with Auto-Greetings)", 1, 15, C["dark_purple"])
    write_title_row(ws, f"Total Events: {len(rows)}  |  Generated: {TODAY}  |  Website: ligiftbasket.com  |  Location: Plainview, Long Island, NY", 2, 15, C["purple"], C["white"], 10)
    write_header_row(ws, HEADERS_MASTER, row=3, bg=C["dark_purple"])

    for i, row in enumerate(rows, 4):
        data = [
            row["Date"], row["Day"], row["Month"], row["Event"], row["Category"],
            row["Emoji"], row["Discount %"], row["Days Until"], row["Status"], row["Alert"],
            row["Promo Code"], row["Best Basket Idea"],
            row["Email Subject"], row["Email Body"], row["WhatsApp Message"]
        ]
        # Color by status
        if row["Status"] == "TODAY":
            fill_color = C["gold"]
        elif row["Alert"] == "URGENT (≤7d)":
            fill_color = C["light_red"]
        elif row["Alert"] == "SOON (≤30d)":
            fill_color = C["light_orange"]
        elif row["Status"] == "PAST":
            fill_color = "F2F3F4"
        else:
            fill_color = C["light_gray"] if i % 2 == 0 else C["white"]

        for c, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = _font(size=9, color="555555" if row["Status"] == "PAST" else "000000")
            cell.fill = _fill(fill_color)
            cell.alignment = _align(wrap=True)
            cell.border = _border(color="DDDDDD")

        # Row height based on body length
        body_len = len(str(row.get("Email Body", "")))
        ws.row_dimensions[i].height = min(120, max(18, body_len // 6))

    auto_col_widths(ws, MASTER_WIDTHS)
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = True


def build_category_sheet(ws, rows, category, title, header_color, row_color):
    cat_rows = [r for r in rows if r["_cat"] == category]
    upcoming = [r for r in cat_rows if r["Status"] in ("UPCOMING", "TODAY")]
    past = [r for r in cat_rows if r["Status"] == "PAST"]

    write_title_row(ws, title, 1, 12, header_color, C["white"], 13)
    write_title_row(ws,
        f"Total: {len(cat_rows)} events  |  Upcoming: {len(upcoming)}  |  Past: {len(past)}",
        2, 12, row_color, "333333", 10)

    headers = ["Date", "Day", "Event / Occasion", "Emoji", "Discount %",
               "Days Until", "Status", "Promo Code", "Best Basket Idea",
               "Email Subject", "Email Body", "WhatsApp Message"]
    write_header_row(ws, headers, row=3, bg=header_color)

    for i, row in enumerate(cat_rows, 4):
        if row["Status"] == "TODAY":
            bg = C["gold"]
        elif row["Alert"] == "URGENT (≤7d)":
            bg = C["light_red"]
        elif row["Alert"] == "SOON (≤30d)":
            bg = C["light_orange"]
        elif row["Status"] == "PAST":
            bg = "F2F3F4"
        else:
            bg = row_color if i % 2 == 0 else C["white"]

        data = [
            row["Date"], row["Day"], row["Event"], row["Emoji"],
            row["Discount %"], row["Days Until"], row["Status"],
            row["Promo Code"], row["Best Basket Idea"],
            row["Email Subject"], row["Email Body"], row["WhatsApp Message"]
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = _font(size=9, color="666666" if row["Status"] == "PAST" else "111111")
            cell.fill = _fill(bg)
            cell.alignment = _align(wrap=True)
            cell.border = _border(color="DDDDDD")
        ws.row_dimensions[i].height = min(100, max(18, len(str(row.get("Email Body", ""))) // 6))

    # Column widths
    for col, w in [(1,14),(2,11),(3,32),(4,6),(5,10),(6,10),(7,12),(8,16),(9,28),(10,38),(11,55),(12,55)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"


def build_alert_sheet(ws, rows):
    """7-day alert sheet — events happening this week."""
    alert_rows = [r for r in rows if isinstance(r["Days Until"], int) and 0 <= r["Days Until"] <= 7]
    alert_rows.sort(key=lambda x: x["Days Until"])

    write_title_row(ws, f"URGENT — Events Within 7 Days (Send Greetings NOW!)", 1, 12, C["red"], C["white"], 13)
    write_title_row(ws, f"Today: {TODAY}  |  {len(alert_rows)} events need attention this week!", 2, 12, C["light_red"], "333333", 10)

    if not alert_rows:
        ws.cell(row=4, column=1).value = "No events in the next 7 days — you're all caught up!"
        ws.cell(row=4, column=1).font = _font(bold=True, size=12, color=C["dark_green"])
        return

    headers = ["Days Until", "Date", "Event", "Category", "Discount %",
               "Promo Code", "Email Subject", "Email Body (SEND NOW)", "WhatsApp (SEND NOW)"]
    write_header_row(ws, headers, row=3, bg=C["red"])

    for i, row in enumerate(alert_rows, 4):
        bg = C["gold"] if row["Days Until"] == 0 else C["light_red"]
        data = [
            f"{'TODAY!' if row['Days Until']==0 else str(row['Days Until'])+' days'}",
            row["Date"], row["Event"], row["Category"],
            f"{row['Discount %']}%", row["Promo Code"],
            row["Email Subject"], row["Email Body"], row["WhatsApp Message"]
        ]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = _font(bold=(row["Days Until"] == 0), size=9)
            cell.fill = _fill(bg)
            cell.alignment = _align(wrap=True)
            cell.border = _border(color="FFAAAA")
        ws.row_dimensions[i].height = min(120, max(25, len(str(row.get("Email Body", ""))) // 5))

    for col, w in [(1,12),(2,14),(3,30),(4,20),(5,10),(6,16),(7,40),(8,65),(9,65)]:
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"


def build_monthly_summary_sheet(ws, rows):
    """One row per month showing event count and upcoming celebrations."""
    write_title_row(ws, "Monthly Summary — All Celebration Events by Month", 1, 7, C["dark_blue"], C["white"])
    write_header_row(ws, ["Month", "Total Events", "Upcoming", "Major Events This Month",
                           "Top Discount", "Best Action"], row=2, bg=C["dark_blue"])

    months = ["January","February","March","April","May","June",
              "July","August","September","October","November","December"]

    for i, month in enumerate(months, 3):
        month_rows = [r for r in rows if r["Month"] == month]
        upcoming = [r for r in month_rows if r["Status"] in ("UPCOMING","TODAY")]
        top_events = ", ".join(r["Event"] for r in month_rows[:3])
        max_disc = max((r["Discount %"] for r in month_rows), default=0)
        action = f"Send {len(upcoming)} greeting(s)" if upcoming else "No upcoming events"

        bg = C["light_blue"] if i % 2 == 0 else C["white"]
        data = [month, len(month_rows), len(upcoming), top_events, f"{max_disc}%", action]
        for c, val in enumerate(data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = _font(size=10, bold=(c == 1))
            cell.fill = _fill(bg)
            cell.alignment = _align(wrap=True)
            cell.border = _border()
        ws.row_dimensions[i].height = 22

    for col, w in [(1,14),(2,14),(3,12),(4,60),(5,12),(6,30)]:
        ws.column_dimensions[get_column_letter(col)].width = w


# ── MAIN BUILD FUNCTION ───────────────────────────────────────────────────────
def build_workbook():
    print(f"\n{'='*65}")
    print(f"  BUILDING MASTER CELEBRATION & SPECIAL DAYS EXCEL")
    print(f"  Date: {TODAY} | Year: {YEAR}")
    print(f"{'='*65}")

    rows = build_all_days()
    print(f"\n  Total events built: {len(rows)}")
    upcoming = [r for r in rows if r["Status"] in ("UPCOMING", "TODAY")]
    urgent = [r for r in rows if r["Alert"] == "URGENT (≤7d)"]
    print(f"  Upcoming this year: {len(upcoming)}")
    print(f"  Urgent (within 7 days): {len(urgent)}")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Sheet 1: 7-Day Alert (most important — shown first) ──
    print("  Building: 7-Day Alert sheet...")
    ws_alert = wb.create_sheet("URGENT - 7 Day Alerts")
    build_alert_sheet(ws_alert, rows)

    # ── Sheet 2: Master All Days ──
    print("  Building: Master All Days sheet...")
    ws_master = wb.create_sheet("ALL DAYS (Master)")
    build_master_sheet(ws_master, rows)

    # ── Sheet 3: Monthly Summary ──
    print("  Building: Monthly Summary sheet...")
    ws_monthly = wb.create_sheet("Monthly Summary")
    build_monthly_summary_sheet(ws_monthly, rows)

    # ── Category Sheets ──
    categories = [
        ("Graduation & Ceremony",    "GRADUATION & CEREMONIES",          C["dark_purple"], C["light_purple"]),
        ("Major US Holiday",         "MAJOR US HOLIDAYS",                 C["dark_blue"],   C["light_blue"]),
        ("Personal Milestone",       "PERSONAL MILESTONES (Birthdays etc)",C["red"],        C["light_red"]),
        ("Religious",                "RELIGIOUS & CULTURAL",              C["gold"],        C["yellow"]),
        ("Corporate & Professional", "CORPORATE & PROFESSIONAL",          C["dark_gray"],   C["light_gray"]),
        ("Seasonal",                 "SEASONAL EVENTS",                   C["dark_green"],  C["light_green"]),
        ("Long Island Special",      "LONG ISLAND SPECIAL EVENTS",        C["orange"],      C["light_orange"]),
        ("Health & Awareness",       "HEALTH & AWARENESS",                C["blue"],        C["light_blue"]),
        ("Love & Romance",           "LOVE & ROMANCE",                    C["red"],         C["light_red"]),
        ("Food & Fun",               "FOOD & FUN DAYS",                   C["purple"],      C["light_purple"]),
        ("Shopping & Sale",          "SHOPPING & SALE DAYS",              C["maroon"],      C["yellow"]),
    ]

    for cat_key, sheet_title, hc, rc in categories:
        cat_count = len([r for r in rows if r["_cat"] == cat_key])
        print(f"  Building: {sheet_title} ({cat_count} events)...")
        ws = wb.create_sheet(sheet_title[:31])
        build_category_sheet(ws, rows, cat_key, sheet_title, hc, rc)

    # ── Save ──
    out_path = OUTPUT_DIR / f"MASTER_Celebration_SpecialDays_{YEAR}.xlsx"
    wb.save(out_path)

    print(f"\n{'='*65}")
    print(f"  DONE!")
    print(f"  File: {out_path}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"\n  SHEET LIST:")
    for i, name in enumerate(wb.sheetnames, 1):
        print(f"    {i:2d}. {name}")
    print(f"\n  Total events: {len(rows)}")
    print(f"  Upcoming: {len(upcoming)}")
    print(f"  Urgent this week: {len(urgent)}")
    print(f"{'='*65}")
    return out_path


if __name__ == "__main__":
    build_workbook()
